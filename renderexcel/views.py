from django.shortcuts import render, HttpResponse
from .models import dato
from django.views.generic.base import TemplateView
from openpyxl import Workbook
# Create your views here.

def home(request):
    queryset = dato.objects.all()
    return render(request,'renderexcel/index.html',{'queryset':queryset})

class ReporteExcel(TemplateView):
    def get(self, request,*args,**kwargs):
        query = dato.objects.all()
        wb = Workbook()
        bandera = True
        cont = 4
        xd=1


        for q in query:

            if bandera:
                ws = wb.active
                ws.title = 'hoja'+str(xd)
                bandera=False
            else:
                ws = wb.create_sheet('hoja'+str(xd))

            ws.cell(row=3, column=2, ).value = "NOMBRE"
            ws.cell(row=3, column=3, ).value = "APELLIDO"
            ws.cell(row=3, column=4, ).value = "CIUDAD"
            ws.cell(row=3, column=5, ).value = "TELEFONO"

            ws.cell(row=cont, column=2,).value = q.nombre
            ws.cell(row=cont, column=3, ).value = q.apellido
            ws.cell(row=cont, column=4, ).value = q.ciudad
            ws.cell(row=cont, column=5, ).value = q.telefono
            xd += 1



        nombre_archivo = "VasACaerCgupetin.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename= {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response